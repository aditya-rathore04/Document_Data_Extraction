import os
from pathlib import Path
from typing import List, Optional, Union
from PIL import Image
import pymupdf


class LoadedDocument:
    """Represents a loaded document ready for perception / OCR or direct text structuring."""

    def __init__(
        self,
        file_path: Path,
        is_text: bool = False,
        text_content: Optional[str] = None,
        images: Optional[List[Image.Image]] = None,
    ):
        self.file_path = file_path
        self.is_text = is_text
        self.text_content = text_content
        self.images = images or []

    @property
    def page_count(self) -> int:
        if self.is_text:
            return 1
        return len(self.images)

    def __repr__(self) -> str:
        return (
            f"LoadedDocument(path='{self.file_path.name}', is_text={self.is_text}, "
            f"pages={self.page_count})"
        )


def _load_excel_as_markdown(file_path: Path) -> str:
    """Reads Excel workbook (.xlsx, .xls) and converts all non-empty content into clean, compact text tables."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_sections = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            cleaned_rows = []
            for row in rows:
                # Filter out None and empty string cells
                cells = [str(c).strip() for c in row if c is not None and str(c).strip() not in ["", "None"]]
                if cells and not all(c in ["0", "0.0", "None", ""] for c in cells):
                    cleaned_rows.append(" | ".join(cells))

            if not cleaned_rows:
                continue

            section = [f"### Sheet: {sheet_name}"]
            section.extend(cleaned_rows)
            sheet_sections.append("\n".join(section))

        wb.close()
        return "\n\n".join(sheet_sections)
    except Exception as e:
        raise ValueError(f"Failed to read Excel spreadsheet '{file_path.name}': {e}")


def load_document(file_path: Union[str, Path], dpi: int = 120) -> LoadedDocument:
    """
    Loads a PDF, image, Excel spreadsheet, or text file and converts it into standard format:
    - Text/CSV/MD: Reads raw text directly (is_text=True).
    - Excel (.xlsx, .xls): Converts sheets into Markdown tables (is_text=True).
    - PDF: Extracts vector text layer (is_text=True) or renders pages as PIL images for OCR (is_text=False).
    - Image: Opens as PIL Image for OCR (is_text=False).
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Document not found at path: {path}")

    ext = path.suffix.lower()

    # 1. Text, CSV, Markdown, JSON, TSV files
    if ext in [".txt", ".csv", ".tsv", ".md", ".json", ".log"]:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        return LoadedDocument(file_path=path, is_text=True, text_content=content)

    # 2. Excel spreadsheets (.xlsx, .xls)
    if ext in [".xlsx", ".xls", ".xlsm"]:
        excel_text = _load_excel_as_markdown(path)
        return LoadedDocument(file_path=path, is_text=True, text_content=excel_text)

    # 3. PDF Document
    if ext == ".pdf":
        doc = pymupdf.open(path)
        extracted_text_pages: List[str] = []
        images: List[Image.Image] = []
        zoom = dpi / 72.0
        mat = pymupdf.Matrix(zoom, zoom)

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text().strip()
            if text:
                extracted_text_pages.append(f"--- Page {page_num + 1} ---\n{text}")

            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)

        doc.close()

        # If PDF has native selectable text, use it directly (super fast & accurate)
        full_pdf_text = "\n\n".join(extracted_text_pages).strip()
        if len(full_pdf_text) > 30:
            return LoadedDocument(
                file_path=path,
                is_text=True,
                text_content=full_pdf_text,
                images=images,
            )

        # Scanned PDF without text layer: return rendered images for OCR
        return LoadedDocument(file_path=path, is_text=False, images=images)

    # 4. Image file
    if ext in [".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff"]:
        img = Image.open(path).convert("RGB")
        return LoadedDocument(file_path=path, is_text=False, images=[img])

    raise ValueError(f"Unsupported file format '{ext}' for document: {path.name}")
